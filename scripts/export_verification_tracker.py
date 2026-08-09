#!/usr/bin/env python3
"""Export the TowerXchange site-count verification tracker (xlsx).

Read-only over data/gsma.db. Produces a workbook Matt can work through by hand:
what has been extracted so far, what still needs verifying, and where the
highest-value doubts are.

Usage:
    python3 scripts/export_verification_tracker.py [output.xlsx]

Re-run any time to regenerate from the current database. Filled-in progress
lives in the copy you are editing, so save your working copy under a new name
before regenerating.
"""
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
DB = REPO / "data" / "gsma.db"
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else REPO / "TowerXchange-verification-tracker.xlsx"

FONT = "Arial"
INK = "1F2933"
ACCENT = "1F4E79"
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
FILLIN = PatternFill("solid", fgColor="FFFF00")     # cells for Matt to fill
SUBTLE = PatternFill("solid", fgColor="EBF1F8")
WARN = PatternFill("solid", fgColor="FDE9D9")
THIN = Side(style="thin", color="D5DBE1")
BORDER = Border(bottom=THIN)

COUNT_METRIC = "towers"


# ----------------------------------------------------------------- data ----
def fetch():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row

    rows = con.execute("""
        SELECT co.name AS company, co.type AS company_type,
               ct.name AS country, ct.iso3, ct.region,
               o.value, o.segment, o.as_of_year, o.as_of_quarter,
               o.source, o.confidence, o.note
        FROM observations o
        JOIN companies  co ON co.id = o.company_id
        JOIN countries  ct ON ct.id = o.country_id
        WHERE o.deleted = 0 AND o.metric = ?
          AND o.country_id IS NOT NULL AND o.value IS NOT NULL
        ORDER BY o.value DESC
    """, (COUNT_METRIC,)).fetchall()

    # Companies whose league total disagrees with the sum of their per-country
    # counts by >15% — the rows most worth checking first.
    # 6 companies carry more than one league vintage; always take the latest,
    # otherwise the comparison picks an arbitrary year.
    league = con.execute("""
        SELECT co.name, le.towers AS league,
               le.as_of_year AS ly, le.as_of_quarter AS lq,
               (SELECT SUM(o.value) FROM observations o
                 WHERE o.company_id = co.id AND o.metric = ?
                   AND o.country_id IS NOT NULL AND o.deleted = 0
                   AND (o.segment = 'all' OR o.segment IS NULL)) AS persum,
               (SELECT MAX(o.as_of_year * 10 + o.as_of_quarter) FROM observations o
                 WHERE o.company_id = co.id AND o.metric = ?
                   AND o.country_id IS NOT NULL AND o.deleted = 0) AS gvint
        FROM companies co JOIN league_entries le ON le.company_id = co.id
        WHERE le.towers IS NOT NULL
          AND le.id = (SELECT id FROM league_entries WHERE company_id = co.id
                       ORDER BY as_of_year DESC, as_of_quarter DESC, id DESC LIMIT 1)
        ORDER BY le.towers DESC
    """, (COUNT_METRIC, COUNT_METRIC)).fetchall()

    def vint(y, q):
        return f"{y}Q{q}" if y else "unknown"

    disputed, no_breakdown = {}, []
    for r in league:
        lv = vint(r["ly"], r["lq"])
        if r["league"] and r["persum"]:
            gap = (r["league"] - r["persum"]) / max(r["league"], r["persum"])
            if abs(gap) > 0.15:
                gv = r["gvint"]
                gvs = f"{gv // 10}Q{gv % 10}" if gv else "unknown"
                disputed[r["name"]] = (r["league"], r["persum"], gap, lv, gvs)
        elif r["league"]:
            no_breakdown.append((r["name"], r["league"], lv))

    missing_countries = con.execute("""
        SELECT ct.name, ct.iso3, ct.region FROM countries ct
        WHERE ct.id NOT IN (
            SELECT DISTINCT country_id FROM observations
            WHERE deleted = 0 AND metric = ? AND country_id IS NOT NULL)
        ORDER BY ct.name
    """, (COUNT_METRIC,)).fetchall()

    con.close()
    return rows, disputed, no_breakdown, missing_countries


# ---------------------------------------------------------------- style ----
def style_header(ws, row, headers, widths, fillin_from=None):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
        if fillin_from and i >= fillin_from:
            c.fill = PatternFill("solid", fgColor="B8860B")
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def put(ws, r, c, v, *, bold=False, size=10, color=INK, fill=None,
        fmt=None, wrap=False, align=None, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if wrap or align:
        cell.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical="top")
    return cell


def title(ws, text, sub=None):
    put(ws, 1, 1, text, bold=True, size=15, color=ACCENT)
    if sub:
        put(ws, 2, 1, sub, size=10, color="52514E", wrap=True)


# ---------------------------------------------------------------- build ----
def build():
    rows, disputed, no_breakdown, missing = fetch()
    wb = Workbook()

    # ============================================== 1. Read me =============
    ws = wb.active
    ws.title = "Read me"
    ws.column_dimensions["A"].width = 108
    title(ws, "TowerXchange site counts — verification tracker")
    r = 3
    blocks = [
        ("What this is", [
            "Every tower site-count figure we have extracted from the TowerXchange regional guides,",
            "laid out so you can check them by hand against the source PDFs, one row at a time.",
        ]),
        ("The headline finding", [
            f"EXTRACTION is done: {len(rows):,} site-count figures, covering {len({x['country'] for x in rows})} countries "
            f"and {len({x['company'] for x in rows})} companies.",
            "VERIFICATION has not started: every single figure is currently marked 'public_unverified'.",
            "Nothing has yet been checked by a human against the source document.",
            "That is not a fault — it is simply the next body of work, and this sheet is the instrument for it.",
        ]),
        ("How to use it", [
            "1. Open the 'Verify queue' tab. It is sorted biggest number first, so the most",
            "   consequential figures come first. Work top-down; you do not need to finish it.",
            "2. For each row, open the matching TowerXchange guide (the 'Source' column names it),",
            "   find that country's ownership chart, and look up that company's site count.",
            "3. Fill in the YELLOW columns only. Everything to their left is the extracted record.",
            "4. The 'Progress' figures on the Summary tab update themselves as you fill rows in.",
        ]),
        ("What to put in the yellow columns", [
            "Verified?      —  y  if the guide shows the same number;  n  if it does not;  ?  if you cannot tell",
            "Correct value  —  only if it differs: type the number the guide actually shows",
            "Source used    —  e.g. 'Europe guide p.44' — where you found it",
            "Date checked   —  today's date",
            "Notes          —  anything odd (chart unreadable, company named differently, etc.)",
        ]),
        ("Worked example of a filled row", [
            "Company: Cellnex  |  Country: Spain  |  Site count: 8,000  |  Source: TowerXchange Europe guide",
            "Verified? y   |   Correct value: (leave blank)   |   Source used: Europe guide p.51   |   Date: 2026-07-10",
            "Notes: matches the printed pie label exactly",
        ]),
        ("The other tabs", [
            "Summary          — the numbers at a glance, plus your live progress",
            "Priority doubts  — companies whose global total does NOT match the sum of their country",
            "                   figures. These are the highest-value rows to check first.",
            "Missing data     — countries and companies where we have no figures at all (work not yet possible)",
            "Flagged rows     — figures the extraction itself marked as uncertain, curated, or annotated",
        ]),
        ("Important caveat", [
            "A mismatch between a company's global total and the sum of its country figures is often",
            "LEGITIMATE — the league table and the guides are published at different dates. Treat these",
            "as questions to investigate, not as errors to force into agreement.",
        ]),
    ]
    for head, lines in blocks:
        put(ws, r, 1, head, bold=True, size=11, color=ACCENT)
        r += 1
        for ln in lines:
            put(ws, r, 1, ln, size=10, wrap=True)
            ws.row_dimensions[r].height = 13
            r += 1
        r += 1
    put(ws, r, 1, "Generated from data/gsma.db by scripts/export_verification_tracker.py — re-run to refresh.",
        size=9, color="898781", italic=True)

    # ============================================== 2. Verify queue ========
    wq = wb.create_sheet("Verify queue")
    title(wq, "Verify queue — every extracted site count, biggest first",
          "Fill in the yellow columns only. Rows flagged PRIORITY belong to a company whose totals do not reconcile.")
    hdr = 4
    headers = ["#", "Company", "Country", "Region", "Site count", "Segment",
               "As of", "Source (which guide)", "Extraction flag", "Priority",
               "Verified? (y/n/?)", "Correct value", "Source used", "Date checked", "Notes"]
    widths = [5, 30, 20, 10, 12, 10, 9, 30, 16, 11, 15, 13, 22, 13, 34]
    style_header(wq, hdr, headers, widths, fillin_from=11)

    for i, x in enumerate(rows, start=1):
        rr = hdr + i
        flag = []
        if x["confidence"] in ("inferred", "approx"):
            flag.append(x["confidence"])
        if x["source"] and "curated" in x["source"]:
            flag.append("curated")
        if x["note"]:
            flag.append("note")
        prio = "PRIORITY" if x["company"] in disputed else ""
        as_of = f"{x['as_of_year']}Q{x['as_of_quarter']}" if x["as_of_year"] else "unknown"

        put(wq, rr, 1, i, size=9, color="898781")
        put(wq, rr, 2, (x["company"] or "").strip())
        put(wq, rr, 3, x["country"])
        put(wq, rr, 4, x["region"] or "—")
        put(wq, rr, 5, x["value"], fmt="#,##0")
        put(wq, rr, 6, x["segment"] or "all")
        put(wq, rr, 7, as_of)
        put(wq, rr, 8, x["source"] or "—", size=9)
        put(wq, rr, 9, ", ".join(flag), size=9, color="B45309")
        c = put(wq, rr, 10, prio, size=9, bold=bool(prio), color="C0392B")
        if prio:
            c.fill = WARN
        for col in range(11, 16):
            cell = wq.cell(row=rr, column=col)
            cell.fill = FILLIN
            cell.font = Font(name=FONT, size=10)
        wq.cell(row=rr, column=12).number_format = "#,##0"
        wq.cell(row=rr, column=14).number_format = "yyyy-mm-dd"
    last = hdr + len(rows)
    wq.auto_filter.ref = f"A{hdr}:O{last}"

    # ============================================== 3. Summary =============
    sm = wb.create_sheet("Summary", 1)
    sm.column_dimensions["A"].width = 52
    sm.column_dimensions["B"].width = 18
    sm.column_dimensions["C"].width = 46
    title(sm, "Summary", "Progress figures are live formulas — they update as you fill in the Verify queue.")

    r = 4
    put(sm, r, 1, "WORK COMPLETED SO FAR", bold=True, size=11, color=ACCENT); r += 1
    completed = [
        ("Site-count figures extracted from the guides", len(rows), "one row per company, per country, per segment"),
        ("Countries with at least one figure", len({x["country"] for x in rows}), "out of 140 countries in the database"),
        ("Companies with at least one figure", len({x["company"] for x in rows}), "out of 707 companies in the database"),
        ("Total sites represented", sum(x["value"] for x in rows), "sum of every extracted figure"),
        ("Figures hand-curated during extraction", sum(1 for x in rows if x["source"] and "curated" in x["source"]),
         "charts automation could not resolve; already manually read"),
    ]
    for lbl, val, note in completed:
        put(sm, r, 1, lbl); put(sm, r, 2, val, fmt="#,##0", bold=True)
        put(sm, r, 3, note, size=9, color="898781"); r += 1

    r += 1
    put(sm, r, 1, "WORK YET TO COMPLETE", bold=True, size=11, color=ACCENT); r += 1
    put(sm, r, 1, "Figures verified by hand (Verified? = y)")
    put(sm, r, 2, "=COUNTIF('Verify queue'!K:K,\"y\")", fmt="#,##0", bold=True)
    put(sm, r, 3, "starts at 0 — nothing has been checked yet", size=9, color="898781"); r += 1

    put(sm, r, 1, "Figures still unverified")
    put(sm, r, 2, f"={len(rows)}-COUNTIF('Verify queue'!K:K,\"y\")-COUNTIF('Verify queue'!K:K,\"n\")-COUNTIF('Verify queue'!K:K,\"?\")",
        fmt="#,##0", bold=True); r += 1

    put(sm, r, 1, "Verification progress")
    put(sm, r, 2, f"=IFERROR((COUNTIF('Verify queue'!K:K,\"y\")+COUNTIF('Verify queue'!K:K,\"n\")+COUNTIF('Verify queue'!K:K,\"?\"))/{len(rows)},0)",
        fmt="0.0%", bold=True); r += 1

    put(sm, r, 1, "Figures found WRONG so far (Verified? = n)")
    put(sm, r, 2, "=COUNTIF('Verify queue'!K:K,\"n\")", fmt="#,##0", bold=True, color="C0392B"); r += 1

    r += 1
    put(sm, r, 1, "OUTSTANDING GAPS (no data to verify yet)", bold=True, size=11, color=ACCENT); r += 1
    same_vintage = sum(1 for v in disputed.values() if v[3] == v[4])
    gaps = [
        ("Companies whose totals do not reconcile", len(disputed), "see 'Priority doubts' tab"),
        ("   ...of which BOTH figures claim the same quarter", same_vintage,
         "START HERE — 'different publication dates' cannot explain these"),
        ("Countries with no site counts at all", len(missing), "see 'Missing data' tab"),
        ("League companies with no country breakdown", len(no_breakdown), "global total only; no per-country split"),
        ("Figures the extraction flagged as uncertain",
         sum(1 for x in rows if x["confidence"] in ("inferred", "approx")), "see 'Flagged rows' tab"),
    ]
    for lbl, val, note in gaps:
        put(sm, r, 1, lbl); put(sm, r, 2, val, fmt="#,##0", bold=True)
        put(sm, r, 3, note, size=9, color="898781"); r += 1

    r += 1
    put(sm, r, 1, "Every figure in the database is currently marked 'public_unverified' — that is the "
                  "single verification level in use. Establishing higher levels (e.g. 'company_reported', "
                  "'regulator_confirmed') is a decision still to be made.", size=9, color="52514E", wrap=True)
    sm.row_dimensions[r].height = 30

    # ============================================== 4. Priority doubts =====
    pd_ = wb.create_sheet("Priority doubts")
    title(pd_, "Priority doubts — companies whose totals do not reconcile",
          "The league table's global total disagrees with the sum of that company's country figures by more than 15%. "
          "Often legitimate (different publication dates) — but these are the best places to look for a real error.")
    hdr = 4
    style_header(pd_, hdr,
                 ["Company", "League total (global)", "League as of", "Sum of country figures",
                  "Guides as of", "Gap", "Gap %", "Same vintage?", "Resolved? (y/n)", "Notes"],
                 [30, 18, 12, 20, 12, 12, 9, 15, 15, 34], fillin_from=9)
    # same-vintage disputes first: if both figures claim the same quarter, the
    # "different publication dates" explanation cannot apply, so they are the
    # highest-value rows to investigate. Within each group, biggest first.
    ordered = sorted(disputed.items(),
                     key=lambda kv: (kv[1][3] != kv[1][4], -kv[1][0]))
    for i, (name, (lg, ps, gap, lv, gv)) in enumerate(ordered, start=1):
        rr = hdr + i
        put(pd_, rr, 1, name.strip())
        put(pd_, rr, 2, lg, fmt="#,##0")
        put(pd_, rr, 3, lv, size=9)
        put(pd_, rr, 4, ps, fmt="#,##0")
        put(pd_, rr, 5, gv, size=9)
        put(pd_, rr, 6, f"=B{rr}-D{rr}", fmt="#,##0")
        put(pd_, rr, 7, f"=IFERROR((B{rr}-D{rr})/B{rr},0)", fmt="0%")
        # if the two sources share a vintage, a big gap is much harder to explain away
        same = "yes — investigate" if lv == gv else "no — dates differ"
        c = put(pd_, rr, 8, same, size=9,
                color="C0392B" if lv == gv else "52514E",
                bold=(lv == gv))
        if lv == gv:
            c.fill = WARN
        for col in (9, 10):
            cell = pd_.cell(row=rr, column=col)
            cell.fill = FILLIN
            cell.font = Font(name=FONT, size=10)

    # ============================================== 5. Missing data ========
    md = wb.create_sheet("Missing data")
    title(md, "Missing data — where there is nothing to verify yet",
          "These are not errors in the extraction; they are places the source guides give no ownership split.")
    put(md, 4, 1, f"Countries with no per-country site counts ({len(missing)})", bold=True, size=11, color=ACCENT)
    style_header(md, 5, ["Country", "ISO3", "Region", "Why (to establish)", "Action decided"],
                 [30, 10, 14, 40, 30], fillin_from=5)
    for i, m in enumerate(missing, start=1):
        rr = 5 + i
        put(md, rr, 1, m["name"]); put(md, rr, 2, m["iso3"] or "—")
        put(md, rr, 3, m["region"] or "—")
        put(md, rr, 4, "No ownership chart in the guide, or country not covered", size=9, color="52514E")
        cell = md.cell(row=rr, column=5); cell.fill = FILLIN; cell.font = Font(name=FONT, size=10)

    start = 5 + len(missing) + 3
    put(md, start, 1, f"League companies with a global total but no country breakdown ({len(no_breakdown)})",
        bold=True, size=11, color=ACCENT)
    style_header(md, start + 1, ["Company", "League total (global)", "League as of", "Action decided"],
                 [32, 20, 14, 30], fillin_from=4)
    for i, (name, tot, lv) in enumerate(sorted(no_breakdown, key=lambda t: -(t[1] or 0)), start=1):
        rr = start + 1 + i
        put(md, rr, 1, name.strip()); put(md, rr, 2, tot, fmt="#,##0")
        put(md, rr, 3, lv, size=9)
        cell = md.cell(row=rr, column=4); cell.fill = FILLIN; cell.font = Font(name=FONT, size=10)

    # ============================================== 6. Flagged rows ========
    fl = wb.create_sheet("Flagged rows")
    title(fl, "Flagged rows — figures the extraction itself was unsure about",
          "'inferred' = derived as a residual, not printed. 'approx' = read as a '+N' approximation. "
          "'curated' = a human read the chart during extraction. These deserve checking regardless of size.")
    hdr = 4
    style_header(fl, hdr, ["Company", "Country", "Site count", "Flag", "Source", "Extraction note",
                           "Verified? (y/n)", "Notes"],
                 [30, 20, 12, 12, 30, 46, 15, 30], fillin_from=7)
    flagged = [x for x in rows if x["confidence"] in ("inferred", "approx")
               or (x["source"] and "curated" in x["source"])]
    for i, x in enumerate(flagged, start=1):
        rr = hdr + i
        put(fl, rr, 1, (x["company"] or "").strip()); put(fl, rr, 2, x["country"])
        put(fl, rr, 3, x["value"], fmt="#,##0")
        f = x["confidence"] if x["confidence"] in ("inferred", "approx") else "curated"
        put(fl, rr, 4, f, size=9, color="B45309")
        put(fl, rr, 5, x["source"] or "—", size=9)
        put(fl, rr, 6, (x["note"] or "")[:120], size=9, color="52514E")
        for col in (7, 8):
            cell = fl.cell(row=rr, column=col); cell.fill = FILLIN; cell.font = Font(name=FONT, size=10)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Verify queue rows : {len(rows)}")
    print(f"  Priority doubts   : {len(disputed)}")
    print(f"  Missing countries : {len(missing)}")
    print(f"  No-breakdown cos  : {len(no_breakdown)}")
    print(f"  Flagged rows      : {len(flagged)}")


if __name__ == "__main__":
    build()
