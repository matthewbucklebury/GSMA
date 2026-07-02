#!/usr/bin/env python3
"""Consolidate the raw extractions (League Table.xlsx + regional guide pies)
into a single normalised dataset (data/dataset.json).

Inputs:
  - ../League Table.xlsx           (run from data/; adjust ROOT otherwise)
  - extraction/pages_raw.json      (output of extraction/extract_pies.py)
  - overrides_curated.json         (manual corrections for hard charts)

Output:
  - dataset.json
"""
import json, re, sys, unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = Path(__file__).resolve().parent

GUIDE_PUBLICATION = {   # from guide covers: "Qx YYYY UPDATE"
    "MENA": (2025, 1), "LATAM": (2025, 2), "Europe": (2025, 2),
    "Asia": (2024, 4), "Africa": (2025, 2),
}

# Pages whose country cannot be derived from caption/heading automatically
PAGE_COUNTRY = {
    ("Africa", 28): "Malawi",
    ("Africa", 30): "Namibia",
    ("Africa", 31): "Niger",
    ("Africa", 35): "Rwanda",
}
# Page-level segment tags (Italy p26 is the broadcast-tower pie)
PAGE_SEGMENT = {("Europe", 26): "broadcast"}

COUNTRY_FIX = {
    "Burkino Faso": "Burkina Faso",
    "Cote d’Ivoire": "Côte d'Ivoire",
    "ITALY": "Italy",
    "RUSSIA": "Russia",
    "Congo Brazzaville": "Congo Brazzaville",
    "Bissau": "Guinea-Bissau",
    "Guinea Bissau": "Guinea-Bissau",
    "Congo B": "Congo Brazzaville",
    "Ivory Coast": "Côte d'Ivoire",
    "USA": "United States",
    "UK": "United Kingdom",
    "UAE": "United Arab Emirates",
    "DRC": "DRC",
}

ISO3 = {
 "Algeria":"DZA","Bahrain":"BHR","Egypt":"EGY","Iran":"IRN","Iraq":"IRQ","Jordan":"JOR",
 "Kuwait":"KWT","Lebanon":"LBN","Morocco":"MAR","Oman":"OMN","Pakistan":"PAK","Qatar":"QAT",
 "Saudi Arabia":"SAU","Tunisia":"TUN","United Arab Emirates":"ARE","Libya":"LBY","Syria":"SYR",
 "Yemen":"YEM","Palestine":"PSE","Afghanistan":"AFG",
 "Argentina":"ARG","Bolivia":"BOL","Brazil":"BRA","Chile":"CHL","Colombia":"COL",
 "Costa Rica":"CRI","Ecuador":"ECU","El Salvador":"SLV","Guatemala":"GTM","Mexico":"MEX",
 "Nicaragua":"NIC","Panama":"PAN","Paraguay":"PRY","Peru":"PER","Honduras":"HND",
 "Dominican Republic":"DOM","Uruguay":"URY","Puerto Rico":"PRI","Jamaica":"JAM",
 "Bahamas":"BHS","Barbados":"BRB","British Virgin Islands":"VGB","Cayman Islands":"CYM",
 "French West Indies":"GLP","Canada":"CAN",
 "Austria":"AUT","Belgium":"BEL","Bulgaria":"BGR","Croatia":"HRV","Cyprus":"CYP",
 "Czech Republic":"CZE","Denmark":"DNK","Estonia":"EST","Finland":"FIN","France":"FRA",
 "Germany":"DEU","Greece":"GRC","Hungary":"HUN","Iceland":"ISL","Ireland":"IRL",
 "Italy":"ITA","Kazakhstan":"KAZ","Malta":"MLT","Netherlands":"NLD","Norway":"NOR",
 "Poland":"POL","Portugal":"PRT","Romania":"ROU","Russia":"RUS","Serbia":"SRB",
 "Spain":"ESP","Sweden":"SWE","Switzerland":"CHE","Turkey":"TUR","Ukraine":"UKR",
 "United Kingdom":"GBR","Slovakia":"SVK","Slovenia":"SVN","Montenegro":"MNE",
 "Bosnia & Herzegovina":"BIH","North Macedonia":"MKD","Belarus":"BLR","Luxembourg":"LUX",
 "Latvia":"LVA","Lithuania":"LTU","Albania":"ALB","Kosovo":"XKX","Moldova":"MDA",
 "northern Cyprus":"CYP",
 "Australia":"AUS","Bangladesh":"BGD","Cambodia":"KHM","India":"IND","Indonesia":"IDN",
 "Japan":"JPN","Laos":"LAO","Malaysia":"MYS","Mongolia":"MNG","Myanmar":"MMR",
 "Nepal":"NPL","New Zealand":"NZL","Philippines":"PHL","South Korea":"KOR",
 "Sri Lanka":"LKA","Thailand":"THA","Vietnam":"VNM","China":"CHN","Taiwan":"TWN",
 "Singapore":"SGP","Hong Kong":"HKG","Fiji":"FJI","Papua New Guinea":"PNG",
 "Uzbekistan":"UZB","Kyrgyzstan":"KGZ","Tajikistan":"TJK","Turkmenistan":"TKM","Bhutan":"BTN",
 "Angola":"AGO","Burkina Faso":"BFA","Cameroon":"CMR","Chad":"TCD",
 "Congo Brazzaville":"COG","Côte d'Ivoire":"CIV","DRC":"COD","Ethiopia":"ETH",
 "Gabon":"GAB","Ghana":"GHA","Kenya":"KEN","Madagascar":"MDG","Malawi":"MWI",
 "Mozambique":"MOZ","Namibia":"NAM","Niger":"NER","Nigeria":"NGA","Rwanda":"RWA",
 "Senegal":"SEN","South Africa":"ZAF","Tanzania":"TZA","Uganda":"UGA","Zambia":"ZMB",
 "Zimbabwe":"ZWE","Benin":"BEN","Guinea":"GIN","Guinea-Bissau":"GNB","Liberia":"LBR",
 "Sudan":"SDN","South Sudan":"SSD","Mali":"MLI","Mauritania":"MRT","Togo":"TGO",
 "Sierra Leone":"SLE","Somalia":"SOM","Botswana":"BWA","Lesotho":"LSO","Eswatini":"SWZ",
 "Burundi":"BDI","Eritrea":"ERI","Gambia":"GMB","Equatorial Guinea":"GNQ",
 "United States":"USA","China Tower":"CHN","Seychelles":"SYC","Mauritius":"MUS",
 "Djibouti":"DJI","Central African Republic":"CAF",
}

REGION_OF = {}  # filled from guides; league footprints fall back to lookup below
FALLBACK_REGION = {
 "USA":"North America","United States":"North America","Canada":"North America",
 "China":"Asia","India":"Asia","Russia":"Europe","Australia":"Asia",
}

# Labels that are aggregates / non-company buckets rather than a single owner
AGGREGATE_PAT = re.compile(
    r"(mno captive|estimated unknown|other|rooftop sites?$|^rooftops$|small isps"
    r"|local isps|site typologies|mno owned|third party sites)", re.I)

# Company-type curation for pie labels that cannot be resolved automatically.
TYPE_MAP = {
    # broadcasters / state broadcast infrastructure
    "télédiffusion d’algerie": "broadcaster", "tda": "broadcaster",
    "trt (turkish broadcaster)": "broadcaster", "ors austria": "broadcaster",
    "tvr state broadcast": "broadcaster", "media broadcast group": "broadcaster",
    "rai way": "towerco", "ei towers": "towerco", "2rn": "broadcaster",
    "teracom": "broadcaster",
    # government / public bodies
    "opw": "government", "cie": "government", "esb telecoms": "government",
    "usof": "government",
    # known towercos not exactly matching league names
    "atc europe": "towerco", "deutsche funkturm": "towerco",
    "mts / tower infrastructure company": "towerco",
    "towers infra austria": "towerco", "optimus tower": "towerco",
    "5g synergiewerk": "towerco", "telenor infra": "towerco",
    "telia towers norway": "towerco", "telenor tower sweden": "towerco",
    "telia towers sweden": "towerco", "sunab": "jv-infraco",
    "kölbi (ice group)": "mno", "ice": "mno",
    "hightel": "towerco", "towercast": "towerco", "tdf": "towerco",
    "shared access": "towerco", "ap wireless": "towerco", "hibernian": "towerco",
    "highpoint": "towerco", "towercom": "towerco", "cetin": "towerco",
    "helium towers": "towerco", "deodar (jazz)": "towerco",
    "iranian towers": "jv-infraco", "tawal": "towerco", "latis": "towerco",
    "benyatower": "towerco", "hoi": "towerco",
    "global tower corporation (gtc)": "towerco",
    "golden comunicaciones": "towerco", "tower one wireless": "towerco",
    "neutral networks / southern cross (evengroup)": "towerco",
    "intelli site solutions": "towerco", "iimt": "towerco", "mx towers": "towerco",
    "centennial towers": "towerco", "ufinet": "towerco", "torrecom": "towerco",
    "balitower": "towerco", "ibs": "towerco", "gihon": "towerco",
    "protelindo": "towerco", "mitratel": "towerco", "tower bersama": "towerco",
    "stp": "towerco", "ibst": "towerco", "centratama": "towerco",
    "new towers": "towerco", "service telecom": "towerco", "tele2": "mno",
    "vodafone & turk telecom (universal service project)": "jv-infraco",
    "turkcell (universal service project)": "mno", "turk telecom": "mno",
    "nuran": "towerco", "amn": "towerco", "eastcastle infrastructure": "towerco",
    "minara towers": "towerco", "sonatel": "mno", "free senegal": "mno",
    "expresso": "mno", "camtel": "mno", "nexttel": "mno", "moov africa": "mno",
    "swiftnet": "towerco", "mast services": "towerco",
    "towerco of africa": "towerco", "helios towers": "towerco",
    "ihs towers": "towerco", "american tower": "towerco",
    "frontier tower solutions": "towerco", "hardiman telecommunications": "towerco",
    "smart": "mno", "cootel": "mno", "cellcard": "mno", "metfone": "mno",
    "laosat": "mno", "southeast asia tower": "towerco",
    "jtower": "towerco", "docomo": "mno", "kddi": "mno", "softbank": "mno",
    "rakuten": "mno", "inldt": "towerco",
    "summit towers": "towerco", "kirtonkhola towers": "towerco",
    "ab hightech": "towerco", "confidence towers": "towerco", "datco": "towerco",
    "frontier towers": "towerco", "isoc": "towerco", "isón": "towerco",
    "lcc": "towerco", "miescor": "towerco", "philtower": "towerco",
    "unity": "towerco", "edgepoint": "towerco", "edotco": "towerco",
    "ascend telecom": "towerco", "indus towers": "towerco", "altius": "towerco",
    "brookfield towers (altius)": "towerco", "gtl infrastructure": "towerco",
    "tower vision": "towerco", "applied solar technologies": "towerco",
    "bmit technologies": "towerco", "celcomdigi": "mno", "mec": "mno",
    "ministry of communications": "government", "ock": "towerco",
    "ukrtower": "towerco", "ytl": "mno",
    "engro enfrashare": "towerco", "oman tower company": "towerco",
    "dawiyat": "towerco", "lc commence": "towerco", "ntd": "towerco",
    "bulsatkom": "mno", "cnt ecuador": "mno", "dito": "mno",
    "ethio telecom": "mno", "gabon telecom": "mno", "grammenphone": "mno",
    "ltc": "mno", "magyar telecom": "mno", "nepal telcom": "mno",
    "trimob ukrtelecom": "mno", "viva bolivia": "mno", "zitca": "government",
}

def norm(s):
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()
    return s

def parse_quarter(tag):
    """League 'Last updated' tags like 'Q324', 'Q125 ' -> (2024, 3)."""
    if not tag:
        return None, None
    m = re.search(r"Q([1-4])\s*(\d{2})", str(tag))
    if not m:
        return None, None
    return 2000 + int(m.group(2)), int(m.group(1))

def load_league():
    import openpyxl
    wb = openpyxl.load_workbook(ROOT / "League Table.xlsx", data_only=True)
    ws = wb["League Table"]
    rows = list(ws.iter_rows(values_only=True))
    league = []
    for r in rows[1:]:
        if not r or r[1] is None:
            continue
        rank, company, owners, model, towers, footprint, ccount, updated = (list(r) + [None]*8)[:8]
        if towers is None:
            continue
        year, quarter = parse_quarter(updated)
        countries = []
        for c in re.split(r",", footprint or ""):
            c = c.strip().strip("()").strip()
            if not c:
                continue
            countries.append(COUNTRY_FIX.get(c, c))
        league.append({
            "rank": rank, "company": str(company).strip(),
            "owners": (str(owners).strip() if owners else None),
            "business_model": (str(model).strip() if model else None),
            "towers": int(towers),
            "footprint": countries, "country_count": ccount,
            "as_of_year": year, "as_of_quarter": quarter,
            "as_of_raw": str(updated).strip() if updated else None,
        })
    ws2 = wb["Customer table"]
    tenants = []
    for r in list(ws2.iter_rows(values_only=True))[1:]:
        if not r or r[1] is None:
            continue
        tenants.append({
            "company": str(r[1]).strip(),
            "anchor_tenants": [t.strip() for t in re.split(r",", str(r[4] or "")) if t.strip()],
        })
    return league, tenants

def clean_label(label):
    """Legend labels sometimes embed the value (Ireland) or footnote marks."""
    l = re.sub(r"\s+", " ", label).strip()
    l = re.sub(r"\s*[\d,]+\s*$", "", l)          # trailing embedded value
    l = re.sub(r"\*+$", "", l).strip()           # footnote asterisks
    l = re.sub(r"^\d+\.\s*", "", l)              # numbered prefix
    return l

def split_names(text):
    """Split 'MNOs:' / 'Towerco activity:' strings into names, dropping parentheticals."""
    if not text:
        return []
    text = re.sub(r"\([^)]*\)", "", text)
    parts = re.split(r",| and ", text)
    return [p.strip().strip(".").strip() for p in parts if p.strip().strip(".")]

def main():
    league, tenants = load_league()
    pages = json.load(open(DATA / "extraction" / "pages_raw.json"))
    overrides = {k: v for k, v in json.load(open(DATA / "overrides_curated.json")).items()
                 if not k.startswith("_")}

    league_norm = {norm(l["company"]): l for l in league}

    countries = {}
    holdings = []      # {region, country, company, towers, segment, metric, ...}
    presences = []     # MNO presence per country
    problems = []

    for region, blob in pages.items():
        pub_y, pub_q = GUIDE_PUBLICATION[region]
        merged = {}
        for p in blob["pages"]:
            cname = (p["caption_country"] or p["heading_country"]
                     or PAGE_COUNTRY.get((region, p["page"])))
            if not cname:
                problems.append(f"{region} p{p['page']}: no country identified; skipped")
                continue
            cname = COUNTRY_FIX.get(cname.strip(), cname.strip())
            cname = re.sub(r"\s+", " ", cname)
            m = merged.setdefault(cname, {"stats": {}, "pies": []})
            for k, v in (p["stats"] or {}).items():
                m["stats"].setdefault(k, v)
            if p["pie"]:
                m["pies"].append((p["page"], p["pie"]))

        for cname, m in merged.items():
            key = f"{region}/{cname}"
            iso3 = ISO3.get(cname)
            if not iso3:
                problems.append(f"{key}: no ISO3 mapping")
            st = m["stats"]
            def num(x):
                if not x:
                    return None
                mm = re.match(r"([\d,]+)$", x.replace(" ", ""))
                return int(mm.group(1).replace(",", "")) if mm else None
            countries[cname] = {
                "name": cname, "iso3": iso3, "region": region,
                "towers_total": num(st.get("towers_total")),
                "towers_total_raw": st.get("towers_total"),
                "population_raw": st.get("population"),
                "subscribers_raw": st.get("subscribers"),
                "sims_per_tower": num(st.get("sims_per_tower")),
                "sim_penetration": st.get("sim_penetration"),
                "mnos_raw": st.get("mnos"),
                "towerco_activity_raw": st.get("towerco_activity"),
                "as_of_year": pub_y, "as_of_quarter": pub_q,
                "source": f"TowerXchange {region} guide",
            }
            REGION_OF[cname] = region
            mno_names = split_names(st.get("mnos"))
            for mn in mno_names:
                if 1 < len(mn) < 60:
                    presences.append({"region": region, "country": cname,
                                      "company": mn, "role": "mno"})

            if key in overrides:
                ov = overrides[key]
                for e in ov["entries"]:
                    holdings.append({
                        "region": region, "country": cname,
                        "company": e["company"],
                        "value": e["value"], "metric": "towers",
                        "segment": e.get("segment", "all"),
                        "aggregate": bool(e.get("aggregate")),
                        "confidence": e.get("confidence", "reported"),
                        "note": e.get("note") or ov.get("reason"),
                        "as_of_year": pub_y, "as_of_quarter": pub_q,
                        "source": f"TowerXchange {region} guide (curated)",
                    })
                continue
            for page_no, pie in m["pies"]:
                seg = PAGE_SEGMENT.get((region, page_no), "all")
                metric = "market_share_pct" if pie.get("kind") == "share" else "towers"
                for e in pie["entries"]:
                    if "value" not in e:
                        problems.append(f"{key} p{page_no}: no value for '{e['label']}'")
                        continue
                    label = clean_label(e["label"])
                    if not label:
                        problems.append(f"{key} p{page_no}: empty label value={e.get('value_text')}")
                        continue
                    conf = "reported"
                    if e.get("inferred") or e.get("sliver"):
                        conf = "inferred"
                    if e.get("approx"):
                        conf = "approx"
                    if e.get("frac_ok") is False:
                        note = "Pie slice drawn out of proportion in source; value as printed"
                    else:
                        note = None
                    holdings.append({
                        "region": region, "country": cname, "company": label,
                        "value": e["value"], "metric": metric, "segment": seg,
                        "aggregate": bool(AGGREGATE_PAT.search(label)),
                        "confidence": conf, "note": note,
                        "as_of_year": pub_y, "as_of_quarter": pub_q,
                        "source": f"TowerXchange {region} guide",
                    })

    # company classification
    TYPE_MAP_N = {norm(k): v for k, v in TYPE_MAP.items()}

    def classify(name, country=None):
        n = norm(name)
        if AGGREGATE_PAT.search(name):
            return "aggregate"
        if n in TYPE_MAP_N:
            return TYPE_MAP_N[n]
        for tn, l in league_norm.items():
            if tn and (tn == n or (len(n) > 5 and (n in tn or tn in n))):
                bm = norm(l["business_model"] or "")
                if "jv" in bm:
                    return "jv-infraco"
                return "towerco"
        # MNO presence match
        if country:
            for pres in presences:
                if pres["country"] == country and norm(pres["company"]) and (
                        norm(pres["company"]) in n or n in norm(pres["company"])):
                    return "mno"
        # global MNO name match
        alln = {norm(p["company"]) for p in presences}
        if n in alln:
            return "mno"
        for a in alln:
            if a and len(n) > 3 and (n in a or a in n):
                return "mno"
        return "unknown"

    for h in holdings:
        h["company_type"] = classify(h["company"], h["country"])

    unknown = sorted({h["company"] for h in holdings if h["company_type"] == "unknown"})

    dataset = {
        "meta": {
            "generated_from": ["League Table.xlsx", "TowerXchange regional guides (5 PDFs)"],
            "guide_publications": {k: f"Q{q} {y}" for k, (y, q) in GUIDE_PUBLICATION.items()},
        },
        "league": league,
        "tenants": tenants,
        "countries": sorted(countries.values(), key=lambda c: (c["region"], c["name"])),
        "holdings": holdings,
        "mno_presences": presences,
        "problems": problems,
        "unclassified_companies": unknown,
    }
    out = DATA / "dataset.json"
    json.dump(dataset, open(out, "w"), indent=1, ensure_ascii=False)
    print(f"countries: {len(countries)}  holdings: {len(holdings)}  league: {len(league)}")
    print("problems:", len(problems))
    for p in problems[:30]:
        print("  !", p)
    print("unclassified:", unknown)

if __name__ == "__main__":
    main()
